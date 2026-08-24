"""Export the model recap workbook (and optional JSON for other renderers).

  python scripts/export_model_recap.py                       # xlsx to Desktop
  python scripts/export_model_recap.py --out path.xlsx --json path.json

Rerunnable after every retrain: everything is read fresh from nfl.duckdb
(read-only) plus the tuning grids from docs/model_report.md (the training
run's own record). Sheets: Overview, Gate story, By season, Calibration,
Power ratings, Upcoming week, Tuning grids.
"""

import argparse
import json
import re
import sys
from datetime import date
from pathlib import Path

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

REPORT = ROOT / "docs" / "model_report.md"
HOLDOUT = (2019, 2024)


def per_season(con) -> list[dict]:
    return (
        con.execute(
            """
        SELECT season,
               count(*) AS games,
               round(avg(pow(p_home_win - home_win, 2)), 4) AS model_brier,
               round(avg(pow(market_home_prob - home_win, 2)), 4) AS market_brier,
               round(avg(pow(p_home_win - home_win, 2))
                     - avg(pow(market_home_prob - home_win, 2)), 4) AS gap,
               round(avg(abs(pred_margin - home_margin)), 2) AS margin_mae,
               -- model-vs-spread record where the model disagrees with the
               -- closing line by 3+ points (mirrors backtest.py: home covers
               -- when actual margin beats the spread from home perspective)
               sum(CASE WHEN pred_margin - spread_line >= 3
                             AND home_margin > spread_line THEN 1
                        WHEN spread_line - pred_margin >= 3
                             AND home_margin < spread_line THEN 1
                        ELSE 0 END) AS ats3_wins,
               sum(CASE WHEN abs(pred_margin - spread_line) >= 3
                             AND home_margin != spread_line THEN 1
                        ELSE 0 END) AS ats3_bets
        FROM model_predictions
        WHERE market_home_prob IS NOT NULL
        GROUP BY season ORDER BY season
        """
        )
        .fetchdf()
        .to_dict("records")
    )


def calibration(con) -> list[dict]:
    return (
        con.execute(
            f"""
        SELECT least(floor(p_home_win * 10), 9)::int AS bin,
               count(*) AS n,
               round(avg(p_home_win), 3) AS predicted,
               round(avg(home_win), 3) AS actual,
               round(avg(home_win) - avg(p_home_win), 3) AS gap
        FROM model_predictions
        WHERE season BETWEEN {HOLDOUT[0]} AND {HOLDOUT[1]}
        GROUP BY 1 ORDER BY 1
        """
        )
        .fetchdf()
        .to_dict("records")
    )


def ratings(con) -> list[dict]:
    return (
        con.execute(
            """
        SELECT team, r_off_pass, r_off_rush, r_def_pass, r_def_rush,
               round(r_off_pass + r_off_rush - r_def_pass - r_def_rush, 4) AS net
        FROM model_ratings ORDER BY net DESC
        """
        )
        .fetchdf()
        .to_dict("records")
    )


def params(con) -> dict:
    row = con.execute("SELECT * FROM model_params").fetchdf().to_dict("records")[0]
    row["feature_cols"] = json.loads(row["feature_cols"])
    for k in ("win_coefs", "margin_coefs"):
        row.pop(k, None)  # coefficient vectors are noise for the recap
    return {k: (str(v) if hasattr(v, "isoformat") else v) for k, v in row.items()}


def upcoming(con) -> list[dict]:
    from nfl_analytics.model.predict import predict_game

    games = (
        con.execute(
            """
        SELECT game_id, week, away_team, home_team, spread_line, total_line
        FROM schedules
        WHERE season = (SELECT max(season) FROM schedules)
          AND week = (SELECT min(week) FROM schedules
                      WHERE season = (SELECT max(season) FROM schedules)
                        AND result IS NULL)
          AND result IS NULL
        ORDER BY gameday, gametime
        """
        )
        .fetchdf()
        .to_dict("records")
    )
    out = []
    for g in games:
        try:
            p = predict_game(con, g["home_team"], g["away_team"])
            out.append(
                {
                    "game": f"{g['away_team']} @ {g['home_team']}",
                    "week": g["week"],
                    "p_home_win": round(p["p_home_win"], 3),
                    "pred_margin": round(p["pred_margin"], 1),
                    "d_qb_out": p.get("inputs", {}).get("d_qb_out", 0),
                    "spread_line": g["spread_line"],
                    "total_line": g["total_line"],
                }
            )
        except Exception as e:  # a single unpredictable game must not kill the export
            out.append({"game": f"{g['away_team']} @ {g['home_team']}", "error": str(e)})
    return out


def parse_report() -> dict:
    """Gate table, QB-subset line, holdout block, 2025 OOS, tuning grids —
    from the training run's own report so the recap can't drift from it."""
    md = REPORT.read_text(encoding="utf-8")

    def grid(header: str) -> list[list[str]]:
        # first pipe-table anywhere inside the section (prose may precede it)
        m = re.search(rf"## {header}[^\n]*\n(.*?)(?=\n## |\Z)", md, re.S)
        if not m:
            return []
        m = re.search(r"((?:^\|.+\n)+)", m.group(1), re.M)
        if not m:
            return []
        rows = [
            [c.strip() for c in line.strip().strip("|").split("|")]
            for line in m.group(1).strip().splitlines()
        ]
        return [r for r in rows if not set("".join(r)) <= set("-: ")]

    verdict = ""
    vm = re.search(r"\*\*Verdict:?\*\*:?\s*(.+?)(?:\n\n|\Z)", md, re.S)
    if vm:
        verdict = " ".join(vm.group(1).split())
    return {
        "gate_table": grid("Configuration comparison"),
        "ewma_grid": grid("Tuning grid.*EWMA"),
        "ridge_grid": grid("Tuning grid.*ridge"),
        "verdict": verdict,
        "qb_subset": (
            re.search(r"([^\n]*95[^\n]*qb[^\n]*|[^\n]*d_qb_out ≠ 0[^\n]*)", md, re.I) or [""]
        )[0].strip()
        if re.search(r"95", md)
        else "",
    }


def sheet(wb, title: str, header: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, h in enumerate(header, 1):
        width = max([len(str(h))] + [len(str(r[i - 1])) for r in rows if len(r) >= i])
        ws.column_dimensions[get_column_letter(i)].width = min(width + 3, 44)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--out", default=str(Path.home() / "OneDrive" / "Desktop" / "NFL_model_recap.xlsx")
    )
    ap.add_argument("--json", help="also dump the aggregates as JSON")
    args = ap.parse_args()

    con = duckdb.connect(str(ROOT / "nfl.duckdb"), read_only=True)
    data = {
        "generated": date.today().isoformat(),
        "params": params(con),
        "per_season": per_season(con),
        "calibration": calibration(con),
        "ratings": ratings(con),
        "upcoming": upcoming(con),
        "report": parse_report(),
    }
    con.close()

    p = data["params"]
    wb = Workbook()
    wb.remove(wb.active)

    hold = [s for s in data["per_season"] if HOLDOUT[0] <= s["season"] <= HOLDOUT[1]]
    hold_model = sum(s["model_brier"] * s["games"] for s in hold) / sum(s["games"] for s in hold)
    hold_market = sum(s["market_brier"] * s["games"] for s in hold) / sum(s["games"] for s in hold)
    overview = [
        ["Fitted at", p["fitted_at"]],
        [
            "Ratings source",
            f"{p['ratings_source']} (half_life={p['half_life']}, carryover={p['carryover']})",
        ],
        ["QB availability flag", "shipped" if p.get("qb_flag") else "off"],
        ["Features", ", ".join(p["feature_cols"])],
        ["Training games (walk-forward)", p["train_games"]],
        ["Holdout window", f"{HOLDOUT[0]}-{HOLDOUT[1]}"],
        ["Holdout Brier — model", round(hold_model, 4)],
        ["Holdout Brier — market", round(hold_market, 4)],
        ["Gap to market", round(hold_model - hold_market, 4)],
        ["Holdout log loss — model", p.get("holdout_logloss")],
        [
            "2025 out-of-sample Brier (model / market)",
            next(
                (
                    f"{s['model_brier']} / {s['market_brier']}"
                    for s in data["per_season"]
                    if s["season"] == 2025
                ),
                "—",
            ),
        ],
        ["Verdict", data["report"]["verdict"]],
    ]
    sheet(wb, "Overview", ["What", "Value"], overview)

    gt = data["report"]["gate_table"]
    if gt:
        sheet(wb, "Gate story", gt[0], gt[1:])
        if data["report"]["qb_subset"]:
            wb["Gate story"].append([])
            wb["Gate story"].append([data["report"]["qb_subset"]])

    sheet(
        wb,
        "By season",
        [
            "Season",
            "Games",
            "Model Brier",
            "Market Brier",
            "Gap",
            "Margin MAE",
            "ATS wins (3+ pt edge)",
            "ATS bets",
            "ATS win%",
        ],
        [
            [
                s["season"],
                s["games"],
                s["model_brier"],
                s["market_brier"],
                s["gap"],
                s["margin_mae"],
                s["ats3_wins"],
                s["ats3_bets"],
                round(s["ats3_wins"] / s["ats3_bets"], 3) if s["ats3_bets"] else None,
            ]
            for s in data["per_season"]
        ],
    )
    sheet(
        wb,
        "Calibration (holdout)",
        ["Decile bin", "Games", "Predicted home win%", "Actual home win%", "Gap"],
        [[c["bin"], c["n"], c["predicted"], c["actual"], c["gap"]] for c in data["calibration"]],
    )
    sheet(
        wb,
        "Power ratings",
        ["Rank", "Team", "Off pass", "Off rush", "Def pass", "Def rush", "Net"],
        [
            [
                i + 1,
                r["team"],
                round(r["r_off_pass"], 4),
                round(r["r_off_rush"], 4),
                round(r["r_def_pass"], 4),
                round(r["r_def_rush"], 4),
                r["net"],
            ]
            for i, r in enumerate(data["ratings"])
        ],
    )
    sheet(
        wb,
        "Upcoming week",
        ["Game", "Week", "P(home win)", "Pred margin", "QB-out edge", "Spread", "Total"],
        [
            [
                u.get("game"),
                u.get("week"),
                u.get("p_home_win"),
                u.get("pred_margin"),
                u.get("d_qb_out"),
                u.get("spread_line"),
                u.get("total_line"),
            ]
            for u in data["upcoming"]
        ],
    )
    for name, g in (("EWMA", data["report"]["ewma_grid"]), ("Ridge", data["report"]["ridge_grid"])):
        if g:
            sheet(wb, f"Tuning grid — {name}", g[0], g[1:])

    out = Path(args.out)
    wb.save(out)
    print(f"wrote {out}")
    if args.json:
        Path(args.json).write_text(json.dumps(data, indent=1, default=str), encoding="utf-8")
        print(f"wrote {args.json}")


if __name__ == "__main__":
    main()
